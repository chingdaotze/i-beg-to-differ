from pathlib import Path
from typing import (
    Dict,
    Self,
)
from zipfile import ZipFile

from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl.cell.cell import (
    TYPE_STRING,
    TYPE_FORMULA,
    TYPE_NUMERIC,
    TYPE_BOOL,
    # TYPE_NULL,
    TYPE_INLINE,
    TYPE_ERROR,
    TYPE_FORMULA_CACHE_STRING,
)

from i_beg_to_differ.core.data_sources.data_source import DataSource
from i_beg_to_differ.core.extensions.ui import (
    WildcardFilePathInputField,
    WildcardInputField,
)
from i_beg_to_differ.core.base import log_exception
from i_beg_to_differ.core.wildcards_sets import WildcardSets


OPENPYXL_TYPES = {
    TYPE_STRING: 'OPENPYXL_TYPE_STRING',
    TYPE_FORMULA: 'OPENPYXL_TYPE_FORMULA',
    TYPE_NUMERIC: 'OPENPYXL_TYPE_NUMERIC',
    TYPE_BOOL: 'OPENPYXL_TYPE_BOOL',
    # TYPE_NULL: 'OPENPYXL_TYPE_NULL',  # Shares a key with TYPE_NUMERIC
    TYPE_INLINE: 'OPENPYXL_TYPE_INLINE',
    TYPE_ERROR: 'OPENPYXL_TYPE_ERROR',
    TYPE_FORMULA_CACHE_STRING: 'OPENPYXL_TYPE_FORMULA_CACHE_STRING',
}
"""
Openpyxl data types.
"""


class DataSourceExcel(
    DataSource,
):
    """
    Excel file data source.
    """

    path: WildcardFilePathInputField
    """
    Path to the ``*.xlsx`` file.
    """

    sheet: WildcardInputField
    """
    Name of the worksheet within the ``*.xlsx`` file.
    """

    _native_types: Dict[str, str] | None
    """
    Cached native types.
    """

    extension_name = '*.xlsx File'

    def __init__(
        self,
        path: str | None = None,
        sheet: str | None = None,
        description: str | None = None,
        wildcard_sets: WildcardSets | None = None,
    ):

        DataSource.__init__(
            self=self,
            description=description,
        )

        if path is None:
            path = ''

        self.path = WildcardFilePathInputField(
            path=path,
            file_dialog_caption='Open *.xlsx File',
            file_dialog_filter='Office Open XML Workbook (*.xlsx)',
            wildcard_sets=wildcard_sets,
            title='File Path',
        )

        if sheet is None:
            sheet = ''

        self.sheet = WildcardInputField(
            base_value=sheet,
            wildcard_sets=wildcard_sets,
            title='Sheet Name',
        )

        self._native_types = None

    def __str__(
        self,
    ) -> str:

        return f'{self.extension_name}: \'[{self.path.base_value}]{self.sheet.base_value}\''

    def load(
        self,
    ) -> DataFrame:
        path = str(
            self.path,
        )

        self.log_info(
            msg=f'Loading Excel file: {path} ...',
        )

        # Load worksheet
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook[str(self.sheet)]

        # Get data types
        self._native_types = {}

        min_col = worksheet.min_column
        max_col = worksheet.max_column

        min_row = worksheet.min_row
        max_row = worksheet.max_row

        for column in range(min_col, max_col + 1):
            # Get column name
            column_name = str(
                worksheet.cell(
                    row=min_row,
                    column=column,
                ).value,
            )

            # Get column types
            unique_column_types = set(
                OPENPYXL_TYPES[
                    worksheet.cell(
                        row=row,
                        column=column,
                    ).data_type
                ]
                for row in range(
                    min_row + 1,
                    max_row + 1,
                )
            )

            # Identify single column type
            if not unique_column_types:
                unique_column_types = None

            else:
                unique_column_types = '|'.join(unique_column_types)

            self._native_types[column_name] = unique_column_types

        # Create dataframe
        data = worksheet.values
        columns = next(data)
        data = list(data)

        return DataFrame(
            data=data,
            columns=columns,
        )

    @classmethod
    @log_exception
    def deserialize(
        cls,
        instance_data: Dict,
        working_dir_path: Path,
        ib2d_file: ZipFile,
        wildcard_sets: WildcardSets | None = None,
    ) -> Self:

        return DataSourceExcel(
            path=instance_data['parameters']['path'],
            sheet=instance_data['parameters']['sheet'],
            wildcard_sets=wildcard_sets,
        )

    @log_exception
    def serialize(
        self,
        ib2d_file: ZipFile,
    ) -> Dict:

        return {
            'extension_id': self.get_extension_id(),
            'description': self.description.value,
            'parameters': {
                'path': self.path.base_value,
                'sheet': self.sheet.base_value,
            },
        }

    @property
    def native_types(
        self,
    ) -> Dict[str, str]:

        if self._native_types is None:
            self.load()

        return self._native_types
